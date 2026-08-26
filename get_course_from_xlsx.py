# -*- coding: utf-8 -*-

import openpyxl
import re

class CourseScheduleParser:
    """
    课程表解析器

    返回值:
    {
        "list": [                          # 按天排列，list[天索引][节索引] = 课名
            ["数学", "语文", "午休", ...],  # 星期一
            ["语文", "英语", "午休", ...],  # 星期二
            ...
        ],
        "time": [                          # 每节课时间 [开始时, 开始分, 结束时, 结束分]
            [8, 0, 8, 45],                 # 第1节
            [8, 55, 9, 40],                # 第2节
            ...
        ]
    }
    空课程位置为 ""
    """

    DAY_KEYWORDS = {
        '周一': '星期一', '星期一': '星期一', 'Mon': '星期一',
        '周二': '星期二', '星期二': '星期二', 'Tue': '星期二',
        '周三': '星期三', '星期三': '星期三', 'Wed': '星期三',
        '周四': '星期四', '星期四': '星期四', 'Thu': '星期四',
        '周五': '星期五', '星期五': '星期五', 'Fri': '星期五',
        '周六': '星期六', '星期六': '星期六', 'Sat': '星期六',
        '周日': '星期日', '星期天': '星期日', '星期日': '星期日', 'Sun': '星期日', '周天': '星期日',
    }

    DAY_ORDER = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

    SCHEDULE_KEYWORDS = ['节次', '星期', '时段', '时间', '节/星', '课程']
    ACTIVITY_KEYWORDS = ['午休', '眼保健操', '大课间', '课间操', '休息', '自习', '放学', '升旗', '读报']

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ws = None
        self.merged_ranges = []
        self.header_row = None
        self.header_col = None
        self.day_columns = {}
        self.period_rows = []
        self.row_merge_info = {}  # {行号: 活动名} 横向合并

    def _get_cell_value(self, row, col):
        """获取单元格值，自动处理合并单元格（返回左上角值）"""
        cell_val = self.ws.cell(row=row, column=col).value
        if cell_val is not None and str(cell_val).strip() != '':
            return cell_val
        for mr in self.merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return self.ws.cell(row=mr.min_row, column=mr.min_col).value
        return None

    def _is_merged(self, row, col):
        """判断单元格是否在合并区域内"""
        for mr in self.merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return mr
        return None

    def _detect_horizontal_merge(self, row):
        """检测横向合并（全天统一活动，如午休、眼保健操）"""
        if not self.day_columns:
            return None
        day_cols = sorted(self.day_columns.values())
        for mr in self.merged_ranges:
            if (mr.min_row == row and mr.max_row == row and
                mr.min_col <= day_cols[0] and mr.max_col >= day_cols[-1] and
                (mr.max_col - mr.min_col) >= 1):
                return (mr.min_col, mr.max_col,
                        self.ws.cell(row=mr.min_row, column=mr.min_col).value)
        return None

    def _parse_time(self, text):
        """
        从文本中解析时间 → [h1, m1, h2, m2]
        支持: "8:00-8:45", "(8:00-8:45)", 全角冒号"8：00-8：45" 等
        """
        if not text:
            return None
        text = str(text).replace('：', ':')
        m = re.search(r'(\d{1,2}):(\d{2})\s*[-~～—]\s*(\d{1,2}):(\d{2})', text)
        if m:
            return [int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))]
        return None

    def locate_schedule(self):
        """自动定位课程表（遍历所有Sheet，找含"节次"+"星期"的行）"""
        for sheet in self.wb.worksheets:
            merged_ranges = list(sheet.merged_cells.ranges)
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 100), values_only=False):
                row_text = ' '.join([str(c.value).strip() for c in row if c.value])
                has_sched = any(kw in row_text for kw in self.SCHEDULE_KEYWORDS)
                has_day = any(kw in row_text for kw in self.DAY_KEYWORDS.keys())
                if has_sched and has_day:
                    self.ws = sheet
                    self.merged_ranges = merged_ranges
                    self.header_row = row[0].row
                    for cell in row:
                        val = str(cell.value).strip() if cell.value else ''
                        if any(kw in val for kw in ['节次', '时段', '时间', '节/星']):
                            self.header_col = cell.column
                            break
                    if self.header_col is None:
                        self.header_col = 1
                    return True
        return False

    def parse_header(self):
        """解析表头，识别星期几对应哪一列"""
        for col in range(1, self.ws.max_column + 1):
            val = self._get_cell_value(self.header_row, col)
            if val is None:
                continue
            val_str = str(val).strip()
            for keyword, day_name in self.DAY_KEYWORDS.items():
                if keyword in val_str:
                    self.day_columns[day_name] = col
                    break

    def parse_periods(self):
        """解析节次行 + 检测横向合并"""
        for row in range(self.header_row + 1, self.ws.max_row + 1):
            val = self._get_cell_value(row, self.header_col)
            if val is not None:
                val_str = str(val).strip()
                if any(c.isdigit() for c in val_str) or any(kw in val_str for kw in self.ACTIVITY_KEYWORDS):
                    self.period_rows.append(row)
                    # 检测横向合并
                    h_merge = self._detect_horizontal_merge(row)
                    if h_merge:
                        self.row_merge_info[row] = str(h_merge[2]).strip()

    def parse(self):
        """
        解析课程表，返回结构化数据
        {"list": [[课...], ...], "time": [[h,m,h,m], ...]}
        """
        if not self.locate_schedule():
            return None
        self.parse_header()
        self.parse_periods()

        # 按星期排序
        sorted_days = [d for d in self.DAY_ORDER if d in self.day_columns]

        # ---- 解析时间 ----
        time_list = []
        for row in self.period_rows:
            val = self._get_cell_value(row, self.header_col)
            t = self._parse_time(val) if val else None
            time_list.append(t if t else [])

        # ---- 解析课程列表 ----
        course_list = []
        for day_name in sorted_days:
            col = self.day_columns[day_name]
            day_courses = []
            skip_rows_v = set()  # 纵向合并覆盖的行

            for row in self.period_rows:
                # 横向合并行（全天活动：午休、眼保健操等）
                if row in self.row_merge_info:
                    day_courses.append(self.row_merge_info[row])
                    continue

                # 纵向合并覆盖行（连堂延续 → 填入相同课名）
                if row in skip_rows_v:
                    val = self._get_cell_value(row, col)
                    day_courses.append(str(val).strip() if val else "")
                    continue

                # 正常读取
                val = self._get_cell_value(row, col)
                val_str = str(val).strip() if val else ""

                # 检查纵向合并（连堂课）
                merged = self._is_merged(row, col)
                if merged:
                    span_rows = merged.max_row - merged.min_row + 1
                    span_cols = merged.max_col - merged.min_col + 1
                    if span_rows > 1 and span_cols == 1:
                        for r in range(merged.min_row + 1, merged.max_row + 1):
                            if r in self.period_rows:
                                skip_rows_v.add(r)

                day_courses.append(val_str)

            course_list.append(day_courses)

        return {
            "list": course_list,
            "time": time_list
        }


# ============ 使用方式 ============
if __name__ == "__main__":
    parser = CourseScheduleParser('你的课程表.xlsx')
    result = parser.parse()
